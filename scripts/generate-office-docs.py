#!/usr/bin/env python3
"""
generate-office-docs.py — Generate JDS-compliant Excel workbooks.

Generates timesheet, expense report, and mileage log workbooks following
the Johansson Documentation System (JDS) visual standards.

Design principles (JDS-PRO-007):
  - Portrait A4 orientation only
  - White background everywhere (no visible grid)
  - No truncated text — everything readable on first open
  - JDS colour palette: Navy Blue headers, alternating row shading
  - Logo embedded in document header (Johansson Engineering by default)

Usage:
    python3 scripts/generate-office-docs.py timesheet [output.xlsx]
    python3 scripts/generate-office-docs.py expense [output.xlsx]
    python3 scripts/generate-office-docs.py mileage [output.xlsx]
    python3 scripts/generate-office-docs.py all [output-dir]

Dependencies:
    pip3 install openpyxl
"""

import argparse
import os
import sys
from datetime import date, timedelta
from pathlib import Path

try:
    import openpyxl
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.page import PageMargins
except ImportError:
    print("Error: openpyxl is required. Install with: pip3 install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ASSETS_DIR = os.path.join(SCRIPT_DIR, '..', 'jds', 'assets')
LOGO_PATH = os.path.join(ASSETS_DIR, 'logo.png')

# ---------------------------------------------------------------------------
# JDS Colour Palette (JDS-PRO-007 §6.1)
# ---------------------------------------------------------------------------
NAVY_BLUE = "1B3A5C"
STEEL_BLUE = "4A90A4"
WARM_GRAY = "8C8C8C"
FOREST_GREEN = "3D8B6E"
SIGNAL_RED = "D04040"
WHITE = "FFFFFF"
ALT_ROW = "F0F3F6"

# ---------------------------------------------------------------------------
# Reusable Styles
# ---------------------------------------------------------------------------
FONT_HEADER = Font(name="Calibri", size=10, bold=True, color=WHITE)
FONT_BODY = Font(name="Calibri", size=10)
FONT_BODY_BOLD = Font(name="Calibri", size=10, bold=True)
FONT_TITLE = Font(name="Calibri", size=14, bold=True, color=NAVY_BLUE)
FONT_SUBTITLE = Font(name="Calibri", size=11, bold=True, color=NAVY_BLUE)
FONT_META_LABEL = Font(name="Calibri", size=9, bold=True, color=NAVY_BLUE)
FONT_META_VALUE = Font(name="Calibri", size=9, color="333333")
FONT_SUMMARY_LABEL = Font(name="Calibri", size=10, bold=True, color=NAVY_BLUE)
FONT_SUMMARY_VALUE = Font(name="Calibri", size=10, bold=True, color=NAVY_BLUE)
FONT_FOOTER = Font(name="Calibri", size=8, italic=True, color=WARM_GRAY)
FONT_UNCONTROLLED = Font(name="Calibri", size=7, italic=True, color=WARM_GRAY)

FILL_NAVY = PatternFill(start_color=NAVY_BLUE, end_color=NAVY_BLUE, fill_type="solid")
FILL_WHITE = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
FILL_ALT = PatternFill(start_color=ALT_ROW, end_color=ALT_ROW, fill_type="solid")

THIN_SIDE = Side(style="thin", color="D0D0D0")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
NO_BORDER = Border()

ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Helper Functions
# ---------------------------------------------------------------------------
def white_fill_all(ws, max_row=80, max_col=20):
    """Fill every cell in the visible area with white background."""
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).fill = FILL_WHITE


def apply_header_row(ws, row, columns, col_start=1):
    """Apply Navy Blue header styling to a row of column headers."""
    for i, col_name in enumerate(columns):
        cell = ws.cell(row=row, column=col_start + i, value=col_name)
        cell.font = FONT_HEADER
        cell.fill = FILL_NAVY
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER


def apply_body_cell(ws, row, col, value=None, font=None, alignment=None, number_format=None):
    """Apply standard body styling to a cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font or FONT_BODY
    cell.alignment = alignment or ALIGN_LEFT
    cell.border = THIN_BORDER
    cell.fill = FILL_WHITE
    if number_format:
        cell.number_format = number_format
    return cell


def apply_alternating_rows(ws, start_row, end_row, start_col, end_col):
    """Apply alternating white / ALT_ROW shading."""
    for r in range(start_row, end_row + 1):
        fill = FILL_ALT if (r - start_row) % 2 == 1 else FILL_WHITE
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).fill = fill


def set_column_widths(ws, widths):
    """Set explicit column widths. widths = {col_letter: width}"""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_logo(ws, row=1, col=1):
    """Embed the JDS logo in the worksheet if available."""
    logo_path = LOGO_PATH
    # Allow override via environment variable for client logos
    env_logo = os.environ.get("JDS_LOGO_PATH")
    if env_logo and os.path.exists(env_logo):
        logo_path = env_logo

    if not os.path.exists(logo_path):
        return
    try:
        img = XlImage(logo_path)
        # Scale to ~52pt (about 70px) for documents
        img.width = 70
        img.height = 70
        cell_ref = f"{get_column_letter(col)}{row}"
        ws.add_image(img, cell_ref)
    except Exception:
        pass  # Logo is nice-to-have, not critical


def write_header_block(ws, doc_title, fields, logo=True, num_cols=7):
    """
    Write the JDS document header block with optional logo.
    Returns the next available row after the header block.
    """
    row = 1

    # Row 1: Title spanning full width
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    title_cell = ws.cell(row=row, column=1, value=doc_title)
    title_cell.font = FONT_TITLE
    title_cell.alignment = ALIGN_LEFT
    # Force white fill on merged region
    for c in range(1, num_cols + 1):
        ws.cell(row=row, column=c).fill = FILL_WHITE

    # "UNCONTROLLED COPY" in last column area (row 2, outside title merge)
    uc_cell = ws.cell(row=2, column=num_cols, value="UNCONTROLLED COPY")
    uc_cell.font = FONT_UNCONTROLLED
    uc_cell.alignment = Alignment(horizontal="right", vertical="top")

    # Logo (overlaid on top-right area)
    if logo:
        add_logo(ws, row=2, col=num_cols - 1)

    # Metadata fields — label in col A, value in col B-C
    row = 2
    for label, value in fields:
        lbl_cell = ws.cell(row=row, column=1, value=label)
        lbl_cell.font = FONT_META_LABEL
        lbl_cell.alignment = ALIGN_LEFT

        # Merge value across 2 columns for readability
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        val_cell = ws.cell(row=row, column=2, value=value)
        val_cell.font = FONT_META_VALUE
        val_cell.alignment = ALIGN_LEFT
        # White fill
        for c in range(1, num_cols + 1):
            ws.cell(row=row, column=c).fill = FILL_WHITE
        row += 1

    # Blank separator row
    row += 1
    return row


def setup_page(ws):
    """Configure page setup — ALWAYS portrait A4."""
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_margins = PageMargins(
        left=0.5, right=0.5, top=0.5, bottom=0.75, header=0.3, footer=0.3
    )
    ws.oddFooter.left.text = "UNCONTROLLED COPY — Git is the controlled copy"
    ws.oddFooter.left.font_size = 8
    ws.oddFooter.right.text = "Page &P of &N"
    ws.oddFooter.right.font_size = 8
    ws.sheet_properties.pageSetUpPr.fitToPage = False
    # Hide gridlines — white background everywhere
    ws.sheet_view.showGridLines = False


# ---------------------------------------------------------------------------
# Timesheet Generator (JDS-TMP-TSH-001)
# ---------------------------------------------------------------------------
def generate_timesheet(output_path):
    """Generate JDS-TMP-TSH-001 Timesheet workbook — portrait A4."""
    wb = openpyxl.Workbook()

    # --- Sheet 1: Timesheet ---
    ws = wb.active
    ws.title = "Timesheet"
    setup_page(ws)
    white_fill_all(ws, max_row=55, max_col=8)

    today = date.today()
    fields = [
        ("Doc No:", "JDS-TSH-GEN-[NNN]"),
        ("Rev:", "DRAFT"),
        ("Date:", today.strftime("%Y-%m-%d")),
        ("Author:", "[Name]"),
        ("Client:", "[Client name]"),
        ("Project:", "[Project No.]"),
        ("Period:", "[Week / Month YYYY]"),
    ]
    data_start = write_header_block(ws, "Timesheet", fields)

    # Portrait-optimised columns (7 columns — fits A4 portrait)
    columns = [
        "Date", "Day", "Project", "Activity",
        "Hours", "OT", "Notes",
    ]
    apply_header_row(ws, data_start, columns)

    # Explicit column widths for portrait A4 — no truncation
    # Column A must be wide enough for metadata labels ("Period:") = 12 chars
    set_column_widths(ws, {
        'A': 14,   # Date + metadata labels
        'B': 14,   # Day + metadata values
        'C': 14,   # Project + metadata values
        'D': 26,   # Activity (wide — main content)
        'E': 8,    # Hours
        'F': 6,    # OT
        'G': 18,   # Notes
    })

    # Pre-fill 31 rows
    first_of_month = today.replace(day=1)
    day_names = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

    body_start = data_start + 1
    for i in range(31):
        r = body_start + i
        d = first_of_month + timedelta(days=i)
        apply_body_cell(ws, r, 1, d.strftime("%Y-%m-%d"), alignment=ALIGN_CENTER)
        apply_body_cell(ws, r, 2, day_names[d.weekday()], alignment=ALIGN_CENTER)
        apply_body_cell(ws, r, 3, alignment=ALIGN_LEFT)       # Project
        apply_body_cell(ws, r, 4, alignment=ALIGN_LEFT)       # Activity
        apply_body_cell(ws, r, 5, number_format="0.00", alignment=ALIGN_RIGHT)  # Hours
        apply_body_cell(ws, r, 6, number_format="0.00", alignment=ALIGN_RIGHT)  # OT
        apply_body_cell(ws, r, 7, alignment=ALIGN_LEFT)       # Notes

    apply_alternating_rows(ws, body_start, body_start + 30, 1, 7)

    # Summary
    summary_row = body_start + 32
    ws.cell(row=summary_row, column=4, value="Total Hours:").font = FONT_SUMMARY_LABEL
    ws.cell(row=summary_row, column=4).alignment = ALIGN_RIGHT
    total_h = ws.cell(row=summary_row, column=5, value=f"=SUM(E{body_start}:E{body_start+30})")
    total_h.font = FONT_SUMMARY_VALUE
    total_h.number_format = "0.00"
    total_h.alignment = ALIGN_RIGHT
    total_h.border = THIN_BORDER

    ws.cell(row=summary_row + 1, column=4, value="Total Overtime:").font = FONT_SUMMARY_LABEL
    ws.cell(row=summary_row + 1, column=4).alignment = ALIGN_RIGHT
    total_ot = ws.cell(row=summary_row + 1, column=5, value=f"=SUM(F{body_start}:F{body_start+30})")
    total_ot.font = FONT_SUMMARY_VALUE
    total_ot.number_format = "0.00"
    total_ot.alignment = ALIGN_RIGHT
    total_ot.border = THIN_BORDER

    ws.cell(row=summary_row + 2, column=4, value="Grand Total:").font = FONT_SUMMARY_LABEL
    ws.cell(row=summary_row + 2, column=4).alignment = ALIGN_RIGHT
    grand = ws.cell(row=summary_row + 2, column=5, value=f"=E{summary_row}+E{summary_row+1}")
    grand.font = FONT_SUMMARY_VALUE
    grand.number_format = "0.00"
    grand.alignment = ALIGN_RIGHT
    grand.border = THIN_BORDER

    # --- Sheet 2: Summary ---
    ws2 = wb.create_sheet("Summary")
    setup_page(ws2)
    white_fill_all(ws2, max_row=30, max_col=10)

    ws2.cell(row=1, column=1, value="Project Summary").font = FONT_TITLE

    proj_cols = ["Project No.", "Total Hours", "% of Total"]
    apply_header_row(ws2, 3, proj_cols)
    for i in range(10):
        r = 4 + i
        for c in range(1, 4):
            apply_body_cell(ws2, r, c, alignment=ALIGN_CENTER)
    apply_alternating_rows(ws2, 4, 13, 1, 3)

    ws2.cell(row=16, column=1, value="Weekly Summary").font = FONT_TITLE
    week_cols = ["Week", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun", "Total"]
    apply_header_row(ws2, 18, week_cols)
    for i in range(6):
        r = 19 + i
        apply_body_cell(ws2, r, 1, alignment=ALIGN_CENTER)
        for c in range(2, 10):
            cell = apply_body_cell(ws2, r, c, number_format="0.00", alignment=ALIGN_RIGHT)
        ws2.cell(row=r, column=9).value = f"=SUM(B{r}:H{r})"
    apply_alternating_rows(ws2, 19, 24, 1, 9)

    set_column_widths(ws2, {
        'A': 12, 'B': 8, 'C': 8, 'D': 8, 'E': 8, 'F': 8, 'G': 8, 'H': 8, 'I': 10,
    })

    wb.save(output_path)
    print(f"Timesheet saved to: {output_path}")


# ---------------------------------------------------------------------------
# Expense Report Generator (JDS-TMP-EXP-001)
# ---------------------------------------------------------------------------
def generate_expense(output_path):
    """Generate JDS-TMP-EXP-001 Expense Report workbook — portrait A4."""
    wb = openpyxl.Workbook()

    # --- Sheet 1: Expenses ---
    ws = wb.active
    ws.title = "Expenses"
    setup_page(ws)
    white_fill_all(ws, max_row=45, max_col=8)

    today = date.today()
    fields = [
        ("Doc No:", "JDS-EXP-GEN-[NNN]"),
        ("Rev:", "DRAFT"),
        ("Date:", today.strftime("%Y-%m-%d")),
        ("Author:", "[Name]"),
        ("Client:", "[Client name]"),
        ("Project:", "[Project No.]"),
        ("Period:", "[Month YYYY]"),
    ]
    data_start = write_header_block(ws, "Expense Report", fields)

    # 7 columns for portrait A4
    columns = [
        "Date", "Category", "Description",
        "Amount", "VAT %", "VAT", "Total",
    ]
    apply_header_row(ws, data_start, columns)

    set_column_widths(ws, {
        'A': 14,   # Date + metadata labels
        'B': 16,   # Category + metadata values
        'C': 22,   # Description + metadata values
        'D': 12,   # Amount
        'E': 8,    # VAT %
        'F': 10,   # VAT
        'G': 14,   # Total
    })

    # Category dropdown
    categories = "Travel,Accommodation,Meals,Materials,Equipment,Software,Shipping,Other"
    dv = DataValidation(type="list", formula1=f'"{categories}"', allow_blank=True)
    dv.error = "Please select a valid category"
    dv.errorTitle = "Invalid Category"
    ws.add_data_validation(dv)

    body_start = data_start + 1
    num_rows = 20
    for i in range(num_rows):
        r = body_start + i
        apply_body_cell(ws, r, 1, alignment=ALIGN_CENTER)       # Date
        cat_cell = apply_body_cell(ws, r, 2, alignment=ALIGN_LEFT)  # Category
        dv.add(cat_cell)
        apply_body_cell(ws, r, 3, alignment=ALIGN_LEFT)         # Description
        apply_body_cell(ws, r, 4, number_format='#,##0.00', alignment=ALIGN_RIGHT)  # Amount
        apply_body_cell(ws, r, 5, number_format='0%', alignment=ALIGN_CENTER)       # VAT %
        vat_cell = apply_body_cell(ws, r, 6, number_format='#,##0.00', alignment=ALIGN_RIGHT)
        vat_cell.value = f'=IF(D{r}<>"",D{r}*E{r},"")'
        total_cell = apply_body_cell(ws, r, 7, number_format='#,##0.00', alignment=ALIGN_RIGHT)
        total_cell.value = f'=IF(D{r}<>"",D{r}+F{r},"")'

    apply_alternating_rows(ws, body_start, body_start + num_rows - 1, 1, 7)

    # Summary
    last_data = body_start + num_rows - 1
    summary_row = body_start + num_rows + 1

    summaries = [
        ("Subtotal:", f"=SUM(D{body_start}:D{last_data})"),
        ("Total VAT:", f"=SUM(F{body_start}:F{last_data})"),
        ("Grand Total:", f"=SUM(G{body_start}:G{last_data})"),
    ]
    for j, (label, formula) in enumerate(summaries):
        r = summary_row + j
        ws.cell(row=r, column=6, value=label).font = FONT_SUMMARY_LABEL
        ws.cell(row=r, column=6).alignment = ALIGN_RIGHT
        val = ws.cell(row=r, column=7, value=formula)
        val.font = FONT_SUMMARY_VALUE
        val.number_format = "#,##0.00"
        val.alignment = ALIGN_RIGHT
        val.border = THIN_BORDER

    # --- Sheet 2: Category Summary ---
    ws2 = wb.create_sheet("Category Summary")
    setup_page(ws2)
    white_fill_all(ws2, max_row=20, max_col=6)

    ws2.cell(row=1, column=1, value="Category Summary").font = FONT_TITLE

    cat_cols = ["Category", "Count", "Excl. VAT", "VAT", "Incl. VAT"]
    apply_header_row(ws2, 3, cat_cols)

    cat_list = categories.split(",")
    for i, cat in enumerate(cat_list):
        r = 4 + i
        apply_body_cell(ws2, r, 1, cat, alignment=ALIGN_LEFT)
        c_cell = apply_body_cell(ws2, r, 2, alignment=ALIGN_RIGHT)
        c_cell.value = f'=COUNTIF(Expenses!B{body_start}:B{last_data},A{r})'
        c_cell.number_format = "0"
        excl = apply_body_cell(ws2, r, 3, alignment=ALIGN_RIGHT)
        excl.value = f'=SUMIF(Expenses!B{body_start}:B{last_data},A{r},Expenses!D{body_start}:D{last_data})'
        excl.number_format = "#,##0.00"
        vat = apply_body_cell(ws2, r, 4, alignment=ALIGN_RIGHT)
        vat.value = f'=SUMIF(Expenses!B{body_start}:B{last_data},A{r},Expenses!F{body_start}:F{last_data})'
        vat.number_format = "#,##0.00"
        incl = apply_body_cell(ws2, r, 5, alignment=ALIGN_RIGHT)
        incl.value = f'=SUMIF(Expenses!B{body_start}:B{last_data},A{r},Expenses!G{body_start}:G{last_data})'
        incl.number_format = "#,##0.00"

    apply_alternating_rows(ws2, 4, 4 + len(cat_list) - 1, 1, 5)
    set_column_widths(ws2, {'A': 18, 'B': 8, 'C': 14, 'D': 10, 'E': 14})

    wb.save(output_path)
    print(f"Expense report saved to: {output_path}")


# ---------------------------------------------------------------------------
# Mileage Log Generator (JDS-TMP-EXP-002)
# ---------------------------------------------------------------------------
def generate_mileage(output_path):
    """Generate JDS-TMP-EXP-002 Mileage Log workbook — portrait A4."""
    wb = openpyxl.Workbook()

    # --- Sheet 1: Mileage ---
    ws = wb.active
    ws.title = "Mileage"
    setup_page(ws)
    white_fill_all(ws, max_row=55, max_col=8)

    today = date.today()
    fields = [
        ("Doc No:", "JDS-EXP-GEN-[NNN]"),
        ("Rev:", "DRAFT"),
        ("Date:", today.strftime("%Y-%m-%d")),
        ("Author:", "[Name]"),
        ("Vehicle:", "[Registration No.]"),
        ("Period:", "[Month / Year]"),
    ]
    data_start = write_header_block(ws, "Mileage Log", fields)

    columns = [
        "Date", "From", "To", "Purpose / Project",
        "km", "Rate", "Amount",
    ]
    apply_header_row(ws, data_start, columns)

    set_column_widths(ws, {
        'A': 14,   # Date + metadata labels
        'B': 16,   # From + metadata values
        'C': 16,   # To + metadata values
        'D': 22,   # Purpose
        'E': 8,    # km
        'F': 8,    # Rate
        'G': 14,   # Amount
    })

    body_start = data_start + 1
    num_rows = 30
    for i in range(num_rows):
        r = body_start + i
        apply_body_cell(ws, r, 1, alignment=ALIGN_CENTER)  # Date
        apply_body_cell(ws, r, 2, alignment=ALIGN_LEFT)    # From
        apply_body_cell(ws, r, 3, alignment=ALIGN_LEFT)    # To
        apply_body_cell(ws, r, 4, alignment=ALIGN_LEFT)    # Purpose
        apply_body_cell(ws, r, 5, number_format="0.0", alignment=ALIGN_RIGHT)  # km
        apply_body_cell(ws, r, 6, value=25.00, number_format="0.00", alignment=ALIGN_RIGHT)  # Rate
        amount = apply_body_cell(ws, r, 7, number_format="#,##0.00", alignment=ALIGN_RIGHT)
        amount.value = f'=IF(E{r}<>"",E{r}*F{r},"")'

    apply_alternating_rows(ws, body_start, body_start + num_rows - 1, 1, 7)

    # Summary
    last_data = body_start + num_rows - 1
    summary_row = body_start + num_rows + 1

    ws.cell(row=summary_row, column=4, value="Total Distance:").font = FONT_SUMMARY_LABEL
    ws.cell(row=summary_row, column=4).alignment = ALIGN_RIGHT
    dist = ws.cell(row=summary_row, column=5, value=f"=SUM(E{body_start}:E{last_data})")
    dist.font = FONT_SUMMARY_VALUE
    dist.number_format = "#,##0.0"
    dist.alignment = ALIGN_RIGHT
    dist.border = THIN_BORDER

    ws.cell(row=summary_row, column=6, value="Total Amount:").font = FONT_SUMMARY_LABEL
    ws.cell(row=summary_row, column=6).alignment = ALIGN_RIGHT
    amt = ws.cell(row=summary_row, column=7, value=f"=SUM(G{body_start}:G{last_data})")
    amt.font = FONT_SUMMARY_VALUE
    amt.number_format = "#,##0.00"
    amt.alignment = ALIGN_RIGHT
    amt.border = THIN_BORDER

    # --- Sheet 2: Monthly Summary ---
    ws2 = wb.create_sheet("Monthly Summary")
    setup_page(ws2)
    white_fill_all(ws2, max_row=20, max_col=5)

    ws2.cell(row=1, column=1, value="Monthly Summary").font = FONT_TITLE

    month_cols = ["Month", "Trips", "Distance (km)", "Amount (SEK)"]
    apply_header_row(ws2, 3, month_cols)

    months = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December",
    ]
    for i, month in enumerate(months):
        r = 4 + i
        apply_body_cell(ws2, r, 1, month, alignment=ALIGN_LEFT)
        apply_body_cell(ws2, r, 2, number_format="0", alignment=ALIGN_RIGHT)
        apply_body_cell(ws2, r, 3, number_format="#,##0.0", alignment=ALIGN_RIGHT)
        apply_body_cell(ws2, r, 4, number_format="#,##0.00", alignment=ALIGN_RIGHT)

    apply_alternating_rows(ws2, 4, 15, 1, 4)

    total_r = 16
    ws2.cell(row=total_r, column=1, value="Total").font = FONT_SUMMARY_LABEL
    for c in range(2, 5):
        cell = ws2.cell(row=total_r, column=c, value=f"=SUM({get_column_letter(c)}4:{get_column_letter(c)}15)")
        cell.font = FONT_SUMMARY_VALUE
        cell.number_format = "#,##0.00" if c == 4 else "#,##0.0" if c == 3 else "0"
        cell.alignment = ALIGN_RIGHT
        cell.border = THIN_BORDER

    set_column_widths(ws2, {'A': 14, 'B': 8, 'C': 16, 'D': 16})

    wb.save(output_path)
    print(f"Mileage log saved to: {output_path}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="Generate JDS-compliant Excel workbooks (timesheet, expense, mileage).",
    )
    parser.add_argument(
        "type",
        choices=["timesheet", "expense", "mileage", "all"],
        help="Type of document to generate",
    )
    parser.add_argument(
        "output",
        nargs="?",
        default=None,
        help="Output file path (.xlsx) or directory (for 'all')",
    )
    args = parser.parse_args()

    today_str = date.today().strftime("%Y-%m-%d")

    if args.type == "all":
        out_dir = args.output or "."
        os.makedirs(out_dir, exist_ok=True)
        generate_timesheet(os.path.join(out_dir, f"JDS-TMP-TSH-001_Timesheet_{today_str}.xlsx"))
        generate_expense(os.path.join(out_dir, f"JDS-TMP-EXP-001_Expense-Report_{today_str}.xlsx"))
        generate_mileage(os.path.join(out_dir, f"JDS-TMP-EXP-002_Mileage-Log_{today_str}.xlsx"))
        print(f"\nAll documents generated in: {out_dir}")
    elif args.type == "timesheet":
        output = args.output or f"JDS-TMP-TSH-001_Timesheet_{today_str}.xlsx"
        generate_timesheet(output)
    elif args.type == "expense":
        output = args.output or f"JDS-TMP-EXP-001_Expense-Report_{today_str}.xlsx"
        generate_expense(output)
    elif args.type == "mileage":
        output = args.output or f"JDS-TMP-EXP-002_Mileage-Log_{today_str}.xlsx"
        generate_mileage(output)


if __name__ == "__main__":
    main()
